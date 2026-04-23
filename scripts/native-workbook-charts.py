#!/usr/bin/env python3

import json
import math
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def normalize_text(value):
    text = "" if value is None else str(value)
    normalized = unicodedata.normalize("NFD", text)
    stripped = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    lowered = stripped.lower()
    cleaned = "".join(char if char.isalnum() or char.isspace() else " " for char in lowered)
    return " ".join(cleaned.split())


def is_number(value):
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return math.isfinite(value)
    return False


def normalize_chart_type(value):
    return (value or "").strip().lower()


def normalize_color(value):
    if not value:
        return None
    text = str(value).strip().lstrip("#").upper()
    if len(text) == 6 and all(char in "0123456789ABCDEF" for char in text):
        return text
    return None


def legend_position_to_excel(value):
    mapping = {
        "top": "t",
        "right": "r",
        "bottom": "b",
        "left": "l",
    }
    return mapping.get((value or "").strip().lower())


def marker_symbol_to_excel(value):
    mapping = {
        "circle": "circle",
        "square": "square",
        "diamond": "diamond",
        "triangle": "triangle",
        "x": "x",
    }
    return mapping.get((value or "").strip().lower())


def set_shape_fill(graphical_props, fill_color):
    if not graphical_props or not fill_color:
        return
    try:
        graphical_props.solidFill = fill_color
    except Exception:
        pass


def set_shape_line(graphical_props, line_color):
    if not graphical_props or not line_color:
        return
    try:
        if graphical_props.ln is None:
            graphical_props.ln = LineProperties()
        graphical_props.line.solidFill = line_color
    except Exception:
        try:
            graphical_props.ln = LineProperties(solidFill=line_color)
        except Exception:
            pass


def ensure_graphical_props(target):
    if target is None:
        return None

    for attr_name in ("graphical_properties", "graphicalProperties", "spPr"):
        if not hasattr(target, attr_name):
            continue
        existing = getattr(target, attr_name)
        if existing is None:
            existing = GraphicalProperties()
            setattr(target, attr_name, existing)
        return existing

    return None


def to_excel_line_width(value):
    if value is None:
        return None
    try:
        return max(1, int(float(value) * 12700))
    except Exception:
        return None


def build_chart(chart_type, title, x_axis_label, y_axis_label):
    normalized = normalize_chart_type(chart_type)
    if normalized == "horizontal_bar":
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
    elif normalized == "grouped_bar":
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
    elif normalized == "stacked_bar":
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
    elif normalized == "stacked_bar_100":
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "percentStacked"
        chart.overlap = 100
    elif normalized == "line":
        chart = LineChart()
    elif normalized == "area":
        chart = AreaChart()
    elif normalized == "pie":
        chart = PieChart()
        chart.varyColors = True
    elif normalized == "doughnut":
        chart = DoughnutChart()
        chart.holeSize = 55
        chart.varyColors = True
    elif normalized == "scatter":
        chart = ScatterChart()
        chart.scatterStyle = "marker"
    else:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"

    chart.width = 11.5
    chart.height = 6.8
    chart.title = title or "Basquio chart"

    if normalized not in {"pie", "doughnut"}:
        try:
            chart.x_axis.title = x_axis_label or ""
            chart.y_axis.title = y_axis_label or ""
        except Exception:
            pass

    try:
        chart.legend.position = "b"
    except Exception:
        pass

    return chart


def create_or_refresh_readme_sheet(workbook, workbook_formats, charts):
    if "README" in workbook.sheetnames:
        sheet = workbook["README"]
        workbook.remove(sheet)

    sheet = workbook.create_sheet("README", 0)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet["A1"] = "Basquio data tables"
    sheet["A1"].font = Font(bold=True, size=16, color="0B0C0C")
    sheet["A2"] = "Sheet"
    sheet["B2"] = "Use"
    sheet["C2"] = "Native chart"
    sheet["D2"] = "Chart title"

    header_fill = PatternFill("solid", fgColor="1A6AFF")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D6D1C4"))
    for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    chart_by_sheet = {}
    for chart in charts:
        sheet_name = chart.get("sheetName")
        if sheet_name and sheet_name not in chart_by_sheet:
            chart_by_sheet[sheet_name] = chart

    row_index = 3
    for sheet_spec in workbook_formats:
        name = sheet_spec.get("sheetName")
        if not name or name == "README":
            continue
        chart_spec = chart_by_sheet.get(name)
        sheet.cell(row_index, 1, value=name)
        sheet.cell(row_index, 2, value="Evidence table with linked native Excel chart" if chart_spec else "Evidence table")
        sheet.cell(row_index, 3, value="Yes" if chart_spec else "No")
        sheet.cell(row_index, 4, value=chart_spec.get("title") if chart_spec else "")
        row_index += 1

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 42
    return sheet


def ensure_table_style(worksheet, table_style_name):
    if worksheet.max_row < 2 or worksheet.max_column < 2:
        return

    ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    if worksheet.tables:
        for name in list(worksheet.tables):
            del worksheet.tables[name]

    display_name = f"tbl_{normalize_text(worksheet.title).replace(' ', '_')[:18] or 'sheet'}"
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style_name or "TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def style_header_row(worksheet, header_fill_color, header_text_color):
    fill = PatternFill("solid", fgColor=header_fill_color or "1A6AFF")
    font = Font(bold=True, color=header_text_color or "FFFFFF")
    border = Border(bottom=Side(style="thin", color="D6D1C4"))
    for column_index in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(1, column_index)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    worksheet.row_dimensions[1].height = 22


def set_sheet_column_widths(worksheet, columns):
    header_to_spec = {column.get("header"): column for column in columns}
    headers = [worksheet.cell(1, index).value for index in range(1, worksheet.max_column + 1)]
    for index, header in enumerate(headers, start=1):
        spec = header_to_spec.get(header)
        if spec and spec.get("widthChars"):
            worksheet.column_dimensions[get_column_letter(index)].width = float(spec.get("widthChars"))
        elif index == 1:
            worksheet.column_dimensions[get_column_letter(index)].width = 22
        else:
            worksheet.column_dimensions[get_column_letter(index)].width = 14


def build_right_panel_anchor(panel_start_col, panel_start_row, width_cols, height_rows):
    anchor = TwoCellAnchor()
    anchor._from.col = panel_start_col - 1
    anchor._from.row = panel_start_row - 1
    anchor.to.col = panel_start_col + width_cols - 2
    anchor.to.row = panel_start_row + height_rows - 2
    return anchor


def should_enable_data_labels(chart_type, matched_rows, series_count):
    normalized = normalize_chart_type(chart_type)
    if normalized in {"pie", "doughnut"}:
        return True
    if normalized in {"scatter", "line", "area"}:
        return False
    return matched_rows <= 8 and series_count <= 1


def apply_workbook_formats(workbook, workbook_formats):
    for sheet_spec in workbook_formats:
        sheet_name = sheet_spec.get("sheetName")
        if not sheet_name or sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        worksheet.sheet_view.showGridLines = bool(sheet_spec.get("showGridLines")) if sheet_spec.get("showGridLines") is not None else False
        worksheet.freeze_panes = sheet_spec.get("freezePane") or "B2"

        style_header_row(
            worksheet,
            normalize_color(sheet_spec.get("headerFillColor")) or "1A6AFF",
            normalize_color(sheet_spec.get("headerTextColor")) or "FFFFFF",
        )

        headers = [worksheet.cell(1, index).value for index in range(1, worksheet.max_column + 1)]
        header_to_column = {str(header): index + 1 for index, header in enumerate(headers) if header not in (None, "")}

        for column_spec in sheet_spec.get("columns", []):
            header = column_spec.get("header")
            number_format = column_spec.get("excelNumberFormat")
            column_index = header_to_column.get(header)
            if not column_index or not number_format:
                continue

            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_index, column_index)
                if is_number(cell.value):
                    cell.number_format = number_format

        set_sheet_column_widths(worksheet, sheet_spec.get("columns", []))
        ensure_table_style(worksheet, sheet_spec.get("tableStyleName"))


def apply_chart_presentation(chart, chart_spec, requested_headers, matched_rows):
    presentation = chart_spec.get("presentation") or {}
    legend_position = legend_position_to_excel(presentation.get("legendPosition"))
    if legend_position:
        try:
            chart.legend.position = legend_position
        except Exception:
            pass

    category_axis = presentation.get("categoryAxis") or {}
    value_axis = presentation.get("valueAxis") or {}
    category_axis_format = category_axis.get("numberFormat")
    value_axis_format = value_axis.get("numberFormat")
    data_label_format = presentation.get("dataLabelFormat")
    chart_background = normalize_color(presentation.get("chartBackground"))
    plot_background = normalize_color(presentation.get("plotBackground"))
    gridline_color = normalize_color(presentation.get("gridlineColor"))
    gridline_width = to_excel_line_width(presentation.get("gridlineWidth"))

    chart_graphics = ensure_graphical_props(chart)
    set_shape_fill(chart_graphics, chart_background)

    plot_graphics = ensure_graphical_props(getattr(chart, "plot_area", None))
    set_shape_fill(plot_graphics, plot_background)

    if category_axis_format:
        try:
            chart.x_axis.numFmt = category_axis_format
        except Exception:
            pass

    if value_axis_format:
        try:
            chart.y_axis.numFmt = value_axis_format
        except Exception:
            pass

    for axis in [getattr(chart, "x_axis", None), getattr(chart, "y_axis", None)]:
        if axis is None or not getattr(axis, "majorGridlines", None):
            continue
        gridline_graphics = ensure_graphical_props(axis.majorGridlines)
        set_shape_line(gridline_graphics, gridline_color)
        if gridline_graphics and gridline_width:
            try:
                if gridline_graphics.ln is None:
                    gridline_graphics.ln = LineProperties()
                gridline_graphics.ln.w = gridline_width
            except Exception:
                pass

    if data_label_format and normalize_chart_type(chart_spec.get("chartType")) in {
        "pie",
        "doughnut",
        "bar",
        "horizontal_bar",
        "grouped_bar",
        "stacked_bar",
        "stacked_bar_100",
        "line",
        "area",
        "scatter",
    }:
        try:
            chart.dLbls = chart.dLbls or DataLabelList()
            chart.dLbls.showVal = should_enable_data_labels(
                chart_spec.get("chartType"),
                matched_rows,
                max(1, len(requested_headers)),
            )
            chart.dLbls.numFmt = data_label_format
            chart.dLbls.showLeaderLines = False
        except Exception:
            pass

    series_styles = presentation.get("series") or []
    for index, series in enumerate(chart.series):
        style = series_styles[index] if index < len(series_styles) else {}
        fill_color = normalize_color(style.get("color"))
        line_color = normalize_color(style.get("lineColor") or style.get("color"))
        set_shape_fill(series.graphicalProperties, fill_color)
        set_shape_line(series.graphicalProperties, line_color)

        if normalize_chart_type(chart_spec.get("chartType")) == "scatter":
            marker_symbol = marker_symbol_to_excel(style.get("markerSymbol"))
            marker_size = style.get("markerSize")
            try:
                if marker_symbol:
                    series.marker.symbol = marker_symbol
                if marker_size:
                    series.marker.size = int(marker_size)
            except Exception:
                pass
            try:
                set_shape_fill(series.marker.graphicalProperties, fill_color)
                set_shape_line(series.marker.graphicalProperties, line_color)
            except Exception:
                pass

        if not fill_color and index < len(requested_headers):
            fallback_color = normalize_color(series_styles[0].get("color")) if series_styles else None
            if fallback_color:
                set_shape_fill(series.graphicalProperties, fallback_color)
                set_shape_line(series.graphicalProperties, fallback_color)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: native-workbook-charts.py <input.xlsx> <spec.json> <output.xlsx>")

    input_path = Path(sys.argv[1])
    spec_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    spec = json.loads(spec_path.read_text())
    workbook = load_workbook(input_path)
    per_sheet_counts = {}
    results = []

    apply_workbook_formats(workbook, spec.get("workbookFormats", []))
    create_or_refresh_readme_sheet(workbook, spec.get("workbookFormats", []), spec.get("charts", []))

    for chart_spec in spec.get("charts", []):
        sheet_name = chart_spec.get("sheetName")
        if not sheet_name or sheet_name not in workbook.sheetnames:
            results.append({
                "chartId": chart_spec.get("chartId"),
                "sheetName": sheet_name,
                "created": False,
                "reason": "sheet_not_found",
            })
            continue

        worksheet = workbook[sheet_name]
        headers = [worksheet.cell(1, index).value for index in range(1, worksheet.max_column + 1)]
        visible_last_col = len(headers)
        header_to_column = {str(header): index + 1 for index, header in enumerate(headers) if header not in (None, "")}
        category_column = 1
        category_header = headers[0] if headers else "Category"
        requested_headers = [header for header in chart_spec.get("selectedHeaders", []) if header in header_to_column]
        categories = chart_spec.get("categories", []) or []
        row_lookup = {}
        for row_index in range(2, worksheet.max_row + 1):
            label = worksheet.cell(row_index, category_column).value
            normalized = normalize_text(label)
            if normalized and normalized not in row_lookup:
                row_lookup[normalized] = row_index

        selected_rows = []
        for category in categories:
            row_index = row_lookup.get(normalize_text(category))
            if row_index:
                selected_rows.append((category, row_index))

        if not selected_rows:
            selected_rows = [
                (worksheet.cell(row_index, category_column).value, row_index)
                for row_index in range(2, worksheet.max_row + 1)
                if str(worksheet.cell(row_index, category_column).value or "").strip()
            ][: min(12, max(0, worksheet.max_row - 1))]

        if not selected_rows:
            results.append({
                "chartId": chart_spec.get("chartId"),
                "sheetName": sheet_name,
                "created": False,
                "reason": "no_rows",
            })
            continue

        if not requested_headers:
            numeric_headers = []
            for column_index, header in enumerate(headers[1:], start=2):
                values = [worksheet.cell(row_index, column_index).value for _, row_index in selected_rows]
                if any(is_number(value) for value in values):
                    numeric_headers.append(str(header))
            requested_headers = numeric_headers[:3]

        chart_type = normalize_chart_type(chart_spec.get("chartType"))
        if chart_type in {"pie", "doughnut"}:
            requested_headers = requested_headers[:1]
        elif chart_type == "scatter":
            requested_headers = requested_headers[:2]
        elif chart_type in {"bar", "horizontal_bar", "line", "area"}:
            requested_headers = requested_headers[:1]
        elif chart_type in {"grouped_bar", "stacked_bar", "stacked_bar_100"}:
            requested_headers = requested_headers[: min(3, len(requested_headers))]

        requested_headers = [header for header in requested_headers if header in header_to_column]
        if len(requested_headers) == 0:
            results.append({
                "chartId": chart_spec.get("chartId"),
                "sheetName": sheet_name,
                "created": False,
                "reason": "no_numeric_headers",
            })
            continue

        sheet_count = per_sheet_counts.get(sheet_name, 0)
        per_sheet_counts[sheet_name] = sheet_count + 1

        workbook_presentation = (chart_spec.get("presentation") or {}).get("workbookPresentation") or {}
        panel_width_cols = int(workbook_presentation.get("chartPanelMinWidthColumns") or 8)
        panel_height_rows = int(workbook_presentation.get("chartPanelMinHeightRows") or 18)
        panel_start_col = max(visible_last_col + 2, 8)
        panel_start_row = 2 + (sheet_count * (panel_height_rows + 2))

        helper_start_col = max(worksheet.max_column + 2, panel_start_col + panel_width_cols + 2)
        helper_start_row = 1 + (sheet_count * (len(selected_rows) + 4))
        helper_header_row = helper_start_row
        helper_data_start_row = helper_start_row + 1

        worksheet.cell(helper_header_row, helper_start_col, value=category_header or "Category")
        for offset, header in enumerate(requested_headers, start=1):
            worksheet.cell(helper_header_row, helper_start_col + offset, value=header)

        for row_offset, (category_value, source_row_index) in enumerate(selected_rows):
            target_row = helper_data_start_row + row_offset
            worksheet.cell(target_row, helper_start_col, value=category_value)
            for offset, header in enumerate(requested_headers, start=1):
                source_column = header_to_column[header]
                source_cell = worksheet.cell(source_row_index, source_column)
                target_cell = worksheet.cell(target_row, helper_start_col + offset, value=source_cell.value)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format

        for column_offset in range(0, len(requested_headers) + 1):
            worksheet.column_dimensions[get_column_letter(helper_start_col + column_offset)].hidden = True

        chart = build_chart(
            chart_type,
            chart_spec.get("title"),
            chart_spec.get("xAxisLabel"),
            chart_spec.get("yAxisLabel"),
        )

        category_ref = Reference(
            worksheet,
            min_col=helper_start_col,
            min_row=helper_data_start_row,
            max_row=helper_data_start_row + len(selected_rows) - 1,
        )

        if chart_type == "scatter" and len(requested_headers) >= 2:
            x_ref = Reference(
                worksheet,
                min_col=helper_start_col + 1,
                min_row=helper_data_start_row,
                max_row=helper_data_start_row + len(selected_rows) - 1,
            )
            y_ref = Reference(
                worksheet,
                min_col=helper_start_col + 2,
                min_row=helper_data_start_row,
                max_row=helper_data_start_row + len(selected_rows) - 1,
            )
            scatter_series_title = (
                ((chart_spec.get("presentation") or {}).get("series") or [{}])[0].get("label")
                or requested_headers[1]
            )
            scatter_series = Series(
                y_ref,
                xvalues=x_ref,
                title=scatter_series_title,
                title_from_data=False,
            )
            chart.series.append(scatter_series)
        else:
            data_ref = Reference(
                worksheet,
                min_col=helper_start_col + 1,
                max_col=helper_start_col + len(requested_headers),
                min_row=helper_header_row,
                max_row=helper_data_start_row + len(selected_rows) - 1,
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(category_ref)

        apply_chart_presentation(chart, chart_spec, requested_headers, len(selected_rows))

        anchor = f"{get_column_letter(panel_start_col)}{panel_start_row}"
        chart.anchor = build_right_panel_anchor(
            panel_start_col,
            panel_start_row,
            panel_width_cols,
            panel_height_rows,
        )
        worksheet.add_chart(chart)

        results.append({
            "chartId": chart_spec.get("chartId"),
            "sheetName": sheet_name,
            "created": True,
            "anchor": anchor,
            "selectedHeaders": requested_headers,
            "matchedRows": len(selected_rows),
        })

    workbook.save(output_path)
    sys.stdout.write(json.dumps({"charts": results}))


if __name__ == "__main__":
    main()
